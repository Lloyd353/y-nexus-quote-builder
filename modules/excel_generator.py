"""
Export Excel du Devis
========================
Génère une version .xlsx du devis, en parallèle du PDF.
 
Argument commercial : le fichier Excel exporté contient une feuille
supplémentaire "Découvrez nos outils" qui présente la suite complète
Y-NEXUS — un client qui reçoit ce fichier et l'explore tombe dessus
naturellement, ce qui est un pont direct vers la boutique Selar.
 
Utilise pandas + openpyxl (moteur d'écriture .xlsx pour pandas).
openpyxl doit être ajouté à requirements.txt.
"""
 
import io
 
import pandas as pd
 
LIEN_BOUTIQUE = "https://selar.com/m/y-nexus-industrial-group1"
EMAIL_GROUPE = "supportpro.ynexus@gmail.com"
NOM_GROUPE = "Y-NEXUS INDUSTRIAL GROUP"
 
# Les 14 outils complets — contrairement à la sidebar qui n'en montre
# que 4 pour rester compacte, l'Excel exporté peut se permettre de
# lister l'intégralité de l'offre, puisque l'utilisateur consulte ce
# fichier hors ligne, sans contrainte d'espace visuel.
CATALOGUE_COMPLET = [
    ("POWER-BALANCE EXPERT", "Équilibrage des phases, facteurs de foisonnement (Ks) et d'utilisation (Ku)", "IEC 60364 / NEC Art. 220"),
    ("CABLE-SECTION SIZER", "Calcul de section de câbles, chutes de tension et courants de court-circuit", "IEC 60364-5-52 / NF C 15-100"),
    ("SUN-GEN DESIGNER", "Dimensionnement complet de parcs solaires photovoltaïques", "IEC 62446"),
    ("LOGIC-FLOW SIZER", "Dimensionnement des actionneurs pneumatiques, hydrauliques et moteurs", "ISO 5598"),
    ("PLC-I/O MANAGER", "Générateur automatique de listes de câblage et d'adressage E/S automates", "IEC 61131-3"),
    ("SAFETY-SIL CALCULATOR", "Évaluation du niveau de performance de sécurité des machines industrielles", "IEC 61508 / ISO 13849"),
    ("BUS-SYNC PRO", "Calcul des longueurs de réseaux industriels et résistances de terminaison", "Modbus / Profibus / CANopen"),
    ("INSTRUMENT-PICKER", "Sélection et calibration des capteurs industriels (Zone ATEX)", "IEC 60079"),
    ("PCB-TRACK SIZER", "Calcul de la largeur des pistes de cuivre sur circuit imprimé", "IPC-2152"),
    ("THERMAL-SHIELD PRO", "Calcul de la dissipation thermique et choix du radiateur optimal", "JEDEC"),
    ("POWER-SUPPLY DESIGNER", "Dimensionnement des alimentations à découpage (Buck, Boost, Flyback)", "IEC 61204"),
    ("LINK-BUDGET MASTER", "Calcul complet du bilan de liaison hertzien", "Normes UIT-R"),
    ("FIBER-LOSS CALCULATOR", "Calcul du budget optique théorique (atténuation, pertes)", "IEC 60793 / TIA-568"),
    ("ANTENNA-ALIGNER PRO", "Calcul des angles d'élévation, azimut et zone de Fresnel", "—"),
]
 
 
def generer_excel_devis(
    infos: dict,
    lignes: list[dict],
    totaux: dict,
    devise: str,
    taux_tva: float,
) -> bytes:
    """
    Construit le classeur Excel complet et retourne ses bytes,
    prêts à être servis via st.download_button.
 
    Le classeur contient 2 feuilles :
        1. "Devis" — le devis à proprement parler, formaté
        2. "Découvrez nos outils" — catalogue promotionnel
    """
    tampon_memoire = io.BytesIO()
 
    date_str = infos["date"].strftime("%d/%m/%Y") if infos.get("date") else "—"
 
    with pd.ExcelWriter(tampon_memoire, engine="openpyxl") as writer:
        # --------------------------------------------------------
        # FEUILLE 1 — DEVIS
        # --------------------------------------------------------
        # On construit le DataFrame des lignes avec la colonne
        # "Prix Total" calculée, exactement comme dans le PDF.
        lignes_export = []
        for index, ligne in enumerate(lignes, start=1):
            quantite = float(ligne.get("Quantité", 0) or 0)
            prix_unitaire = float(ligne.get("Prix unitaire", 0) or 0)
            lignes_export.append(
                {
                    "N°": index,
                    "Désignation": ligne.get("Désignation", ""),
                    "Référence normative": ligne.get("Référence normative", ""),
                    "Quantité": quantite,
                    "Unité": ligne.get("Unité", ""),
                    "Prix unitaire": prix_unitaire,
                    "Prix Total": quantite * prix_unitaire,
                }
            )
        df_devis = pd.DataFrame(lignes_export)
        df_devis.to_excel(writer, sheet_name="Devis", index=False, startrow=6)
 
        feuille_devis = writer.sheets["Devis"]
 
        # En-tête manuel au-dessus du tableau (métadonnées du devis)
        feuille_devis["A1"] = "DEVIS"
        feuille_devis["A2"] = NOM_GROUPE
        feuille_devis["A4"] = f"Concepteur : {infos.get('prenom', '')} {infos.get('nom', '')}".strip()
        feuille_devis["A5"] = f"Client : {infos.get('client', '—')}    |    Date : {date_str}    |    Secteur : {infos.get('secteur', '—')}"
 
        # Lignes de synthèse ajoutées juste après le tableau de données
        ligne_synthese_debut = 6 + len(df_devis) + 3  # +3 lignes d'espace après le tableau
        feuille_devis.cell(row=ligne_synthese_debut, column=6, value="Total HT")
        feuille_devis.cell(row=ligne_synthese_debut, column=7, value=f"{totaux['total_ht']:,.2f} {devise}")
        feuille_devis.cell(row=ligne_synthese_debut + 1, column=6, value=f"TVA ({taux_tva:.2f}%)")
        feuille_devis.cell(row=ligne_synthese_debut + 1, column=7, value=f"{totaux['montant_tva']:,.2f} {devise}")
        feuille_devis.cell(row=ligne_synthese_debut + 2, column=6, value="TOTAL TTC")
        feuille_devis.cell(row=ligne_synthese_debut + 2, column=7, value=f"{totaux['total_ttc']:,.2f} {devise}")
 
        # Filigrane discret en bas de feuille
        ligne_filigrane = ligne_synthese_debut + 5
        feuille_devis.cell(
            row=ligne_filigrane, column=1,
            value=f"Généré via le logiciel d'ingénierie {NOM_GROUPE} — {LIEN_BOUTIQUE}",
        )
 
        _appliquer_mise_en_forme_feuille_devis(feuille_devis, nb_lignes_donnees=len(df_devis))
 
        # --------------------------------------------------------
        # FEUILLE 2 — DÉCOUVREZ NOS OUTILS (promotion)
        # --------------------------------------------------------
        df_catalogue = pd.DataFrame(
            CATALOGUE_COMPLET, columns=["Outil", "Description", "Norme de référence"]
        )
        df_catalogue.to_excel(writer, sheet_name="Découvrez nos outils", index=False, startrow=4)
 
        feuille_catalogue = writer.sheets["Découvrez nos outils"]
        feuille_catalogue["A1"] = f"⚡ {NOM_GROUPE} — Suite complète de 14 outils d'ingénierie"
        feuille_catalogue["A2"] = "Conception de calculateurs pour ingénierie : Électrotechnique, Électronique, Automatisme, Télécommunication."
        feuille_catalogue["A3"] = f"🛒 Boutique : {LIEN_BOUTIQUE}    |    📧 Contact : {EMAIL_GROUPE}"
 
        _appliquer_mise_en_forme_feuille_catalogue(feuille_catalogue, nb_lignes_donnees=len(df_catalogue))
 
    return tampon_memoire.getvalue()
 
 
def _appliquer_mise_en_forme_feuille_devis(feuille, nb_lignes_donnees: int):
    """
    Applique une mise en forme visuelle cohérente avec le thème
    'Industrial Premium' : en-tête bleu profond, largeurs de colonnes
    ajustées, titre en gras.
 
    Import d'openpyxl.styles fait localement (et non en haut du fichier)
    pour garder ce module lisible : ces classes ne servent qu'ici.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
 
    bleu_profond = "1B3A5C"
    police_titre = Font(bold=True, size=16, color=bleu_profond)
    police_entete_tableau = Font(bold=True, color="FFFFFF")
    remplissage_entete = PatternFill(start_color=bleu_profond, end_color=bleu_profond, fill_type="solid")
 
    feuille["A1"].font = police_titre
    feuille["A2"].font = Font(bold=True, size=11)
 
    # Mise en forme de la ligne d'en-tête du tableau (ligne 7 : startrow=6 -> ligne Excel 7)
    ligne_entete_tableau = 7
    for colonne in range(1, 8):  # 7 colonnes : N° à Prix Total
        cellule = feuille.cell(row=ligne_entete_tableau, column=colonne)
        cellule.font = police_entete_tableau
        cellule.fill = remplissage_entete
        cellule.alignment = Alignment(horizontal="center")
 
    # Largeurs de colonnes ajustées pour la lisibilité
    largeurs = {"A": 6, "B": 40, "C": 22, "D": 10, "E": 8, "F": 16, "G": 16}
    for lettre_colonne, largeur in largeurs.items():
        feuille.column_dimensions[lettre_colonne].width = largeur
 
    # Mise en gras de la ligne TOTAL TTC (toujours la 3e ligne de synthèse)
    ligne_total_ttc = 6 + nb_lignes_donnees + 3 + 2
    feuille.cell(row=ligne_total_ttc, column=6).font = Font(bold=True)
    feuille.cell(row=ligne_total_ttc, column=7).font = Font(bold=True, color=bleu_profond, size=12)
 
 
def _appliquer_mise_en_forme_feuille_catalogue(feuille, nb_lignes_donnees: int):
    """Mise en forme de la feuille promotionnelle du catalogue."""
    from openpyxl.styles import Font, PatternFill, Alignment
 
    bleu_profond = "1B3A5C"
 
    feuille["A1"].font = Font(bold=True, size=14, color=bleu_profond)
    feuille["A2"].font = Font(italic=True, size=10)
    feuille["A3"].font = Font(bold=True, size=10)
 
    ligne_entete_tableau = 5
    police_entete = Font(bold=True, color="FFFFFF")
    remplissage_entete = PatternFill(start_color=bleu_profond, end_color=bleu_profond, fill_type="solid")
    for colonne in range(1, 4):  # 3 colonnes : Outil, Description, Norme
        cellule = feuille.cell(row=ligne_entete_tableau, column=colonne)
        cellule.font = police_entete
        cellule.fill = remplissage_entete
 
    largeurs = {"A": 26, "B": 60, "C": 26}
    for lettre_colonne, largeur in largeurs.items():
        feuille.column_dimensions[lettre_colonne].width = largeur
 
    # Retour à la ligne automatique pour la colonne Description (souvent longue)
    for ligne in range(ligne_entete_tableau + 1, ligne_entete_tableau + 1 + nb_lignes_donnees):
        feuille.cell(row=ligne, column=2).alignment = Alignment(wrap_text=True, vertical="top")
 
